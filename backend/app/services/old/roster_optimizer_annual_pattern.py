#!/usr/bin/env python3
"""
Optimizador anual basado en patrón de 4 semanas
"""

import time
from datetime import datetime, timedelta, date
from typing import Dict, List, Any, Set, Tuple
from collections import defaultdict
import calendar
import json
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment


class AnnualPatternOptimizer:
    """
    Optimizador anual que:
    1. Optimiza las primeras 4 semanas
    2. Repite el patrón durante todo el año
    3. Gestiona vacaciones distribuyendo la carga
    """
    
    def __init__(self, client_data: dict):
        """Inicializa el optimizador anual por patrón"""
        self.client_data = client_data
        self.services = client_data.get('services', [])
        
        # Configuración de vacaciones
        self.vacation_weeks_per_year = 3  # 3 semanas de vacaciones al año
        self.backup_factor = 0.15  # 15% de respaldo para vacaciones
        
    def optimize_year(self, year: int) -> Dict[str, Any]:
        """
        Optimiza el año completo basándose en un patrón de 4 semanas
        
        Args:
            year: Año a optimizar
            
        Returns:
            Diccionario con la solución anual
        """
        start_time = time.time()
        
        print(f"\n{'='*80}")
        print(f"OPTIMIZACIÓN ANUAL POR PATRÓN - AÑO {year}")
        print(f"{'='*80}")
        
        # Paso 1: Optimizar las primeras 4 semanas
        print("\n1. Optimizando patrón base (primeras 4 semanas)...")
        base_pattern = self._optimize_base_pattern(year)
        
        if base_pattern['status'] != 'success':
            return {
                'status': 'failed',
                'reason': 'No se pudo optimizar el patrón base de 4 semanas'
            }
        
        print(f"   ✓ Patrón base optimizado: {base_pattern['drivers_used']} conductores")
        
        # Paso 2: Calcular conductores adicionales para vacaciones
        vacation_backup = self._calculate_vacation_backup(base_pattern['drivers_used'])
        total_drivers = base_pattern['drivers_used'] + vacation_backup
        
        print(f"\n2. Conductores para el año:")
        print(f"   - Base (patrón 4 semanas): {base_pattern['drivers_used']}")
        print(f"   - Respaldo vacaciones: {vacation_backup}")
        print(f"   - Total: {total_drivers}")
        
        # Paso 3: Crear pool de conductores
        drivers = self._create_driver_pool(total_drivers)
        
        # Paso 4: Planificar vacaciones anuales
        print(f"\n3. Planificando vacaciones anuales...")
        vacation_schedule = self._plan_annual_vacations(drivers, year)
        
        # Paso 5: Replicar patrón para todo el año
        print(f"\n4. Replicando patrón para 52 semanas...")
        annual_assignments = self._replicate_pattern_annually(
            base_pattern, drivers, vacation_schedule, year
        )
        
        # Paso 6: Calcular métricas
        metrics = self._calculate_annual_metrics(
            annual_assignments, drivers, vacation_schedule
        )
        
        optimization_time = time.time() - start_time
        
        solution = {
            'status': 'success',
            'year': year,
            'drivers_total': total_drivers,
            'drivers_base': base_pattern['drivers_used'],
            'drivers_backup': vacation_backup,
            'base_pattern': base_pattern,
            'annual_assignments': annual_assignments,
            'vacation_schedule': vacation_schedule,
            'metrics': metrics,
            'optimization_time': optimization_time
        }
        
        print(f"\n{'='*60}")
        print("RESUMEN ANUAL")
        print(f"{'='*60}")
        print(f"Conductores totales: {total_drivers}")
        print(f"Semanas cubiertas: 52")
        print(f"Total asignaciones: {metrics['total_assignments']:,}")
        print(f"Tiempo optimización: {optimization_time:.2f}s")
        
        return solution
    
    def _optimize_base_pattern(self, year: int) -> Dict[str, Any]:
        """Optimiza las primeras 4 semanas del año"""
        from app.services.roster_optimizer_grouped import GroupedRosterOptimizer
        
        # Usar enero como mes base (tiene 31 días, cubre más de 4 semanas)
        optimizer = GroupedRosterOptimizer(self.client_data)
        
        # Optimizar enero completo
        solution = optimizer.optimize_month(year, 1)
        
        if solution['status'] == 'success':
            # Extraer solo las primeras 4 semanas (28 días)
            four_weeks_assignments = []
            start_date = date(year, 1, 1)
            end_date = start_date + timedelta(days=27)  # 28 días = 4 semanas
            
            for assignment in solution['assignments']:
                assign_date = datetime.fromisoformat(assignment['date']).date()
                if start_date <= assign_date <= end_date:
                    four_weeks_assignments.append(assignment)
            
            solution['assignments'] = four_weeks_assignments
            solution['weeks'] = 4
            
            # Recalcular métricas para 4 semanas
            total_hours = 0
            for a in four_weeks_assignments:
                # Calcular horas basándose en tiempo de inicio y fin
                if 'hours' in a:
                    total_hours += a['hours']
                elif 'duration_hours' in a:
                    total_hours += a['duration_hours']
                else:
                    # Estimar 8 horas por turno si no hay información
                    total_hours += 8
            
            solution['metrics']['total_hours'] = total_hours
            solution['metrics']['total_shifts'] = len(four_weeks_assignments)
            
            # Asegurar que cada asignación tenga campo 'hours'
            for a in four_weeks_assignments:
                if 'hours' not in a:
                    if 'duration_hours' in a:
                        a['hours'] = a['duration_hours']
                    else:
                        a['hours'] = 8  # Default
        
        return solution
    
    def _calculate_vacation_backup(self, base_drivers: int) -> int:
        """Calcula conductores adicionales necesarios para cubrir vacaciones"""
        # 3 semanas de vacaciones al año = 5.77% del tiempo
        # Pero necesitamos más cobertura porque las vacaciones no son uniformes
        # y queremos mantener el patrón optimizado
        
        # Usamos 15% de respaldo para tener margen
        backup = max(3, int(base_drivers * self.backup_factor))
        
        # Para casos pequeños, asegurar mínimo respaldo
        if base_drivers <= 20:
            backup = max(4, backup)
        
        return backup
    
    def _create_driver_pool(self, total_drivers: int) -> List[Dict]:
        """Crea el pool de conductores"""
        drivers = []
        
        for i in range(1, total_drivers + 1):
            drivers.append({
                'id': i,
                'name': f'Conductor_{i}',
                'is_backup': i > (total_drivers - self._calculate_vacation_backup(total_drivers)),
                'vacation_weeks': [],
                'weekly_pattern': None  # Se asignará del patrón base
            })
        
        return drivers
    
    def _plan_annual_vacations(self, drivers: List[Dict], year: int) -> Dict[int, List[Dict]]:
        """
        Planifica vacaciones distribuyendo uniformemente durante el año
        Cada conductor toma 3 semanas de vacaciones
        """
        vacation_schedule = {}
        total_drivers = len(drivers)
        
        # Distribuir vacaciones uniformemente durante el año
        # Aproximadamente 15% de conductores de vacaciones cada semana
        weeks_in_year = 52
        
        for driver in drivers:
            driver_id = driver['id']
            vacation_weeks = []
            
            # Asignar 3 semanas de vacaciones distribuidas
            # Primera semana: entre semana 1-17
            week1 = ((driver_id - 1) % 17) + 1
            
            # Segunda semana: entre semana 18-34
            week2 = ((driver_id - 1) % 17) + 18
            
            # Tercera semana: entre semana 35-52
            week3 = ((driver_id - 1) % 18) + 35
            
            vacation_weeks = [week1, week2, week3]
            
            vacation_schedule[driver_id] = {
                'driver_id': driver_id,
                'driver_name': driver['name'],
                'weeks': vacation_weeks,
                'total_days': 21  # 3 semanas * 7 días
            }
            
            driver['vacation_weeks'] = vacation_weeks
        
        return vacation_schedule
    
    def _replicate_pattern_annually(self, base_pattern: Dict, drivers: List[Dict],
                                   vacation_schedule: Dict, year: int) -> Dict[str, List]:
        """
        Replica el patrón de 4 semanas durante todo el año
        ajustando por vacaciones
        """
        annual_assignments = {
            'by_week': {},
            'by_driver': defaultdict(list),
            'summary': []
        }
        
        # Mapear conductores del patrón base a los conductores totales
        pattern_assignments = base_pattern['assignments']
        
        # Agrupar asignaciones del patrón por día relativo (1-28)
        pattern_by_day = defaultdict(list)
        start_date = date(year, 1, 1)
        
        for assignment in pattern_assignments:
            assign_date = datetime.fromisoformat(assignment['date']).date()
            day_number = (assign_date - start_date).days + 1  # Día 1-28
            pattern_by_day[day_number].append(assignment)
        
        # Replicar para cada semana del año
        current_date = date(year, 1, 1)
        week_number = 1
        
        while current_date.year == year and week_number <= 52:
            week_start = current_date
            week_end = current_date + timedelta(days=6)
            
            # Obtener conductores disponibles esta semana (no de vacaciones)
            available_drivers = []
            on_vacation = []
            
            for driver in drivers:
                if week_number in driver.get('vacation_weeks', []):
                    on_vacation.append(driver['id'])
                else:
                    available_drivers.append(driver['id'])
            
            # Asignar turnos de la semana basándose en el patrón
            week_assignments = []
            
            for day_offset in range(7):
                day_date = week_start + timedelta(days=day_offset)
                if day_date.year != year:
                    break
                
                # Obtener el día del patrón (ciclo de 28 días)
                pattern_day = ((week_number - 1) * 7 + day_offset) % 28 + 1
                
                # Copiar asignaciones del patrón para este día
                if pattern_day in pattern_by_day:
                    for pattern_assign in pattern_by_day[pattern_day]:
                        # Buscar un conductor disponible con el mismo índice relativo
                        original_driver_id = pattern_assign['driver_id']
                        
                        # Convertir a int si es string
                        if isinstance(original_driver_id, str):
                            # Extraer número del string (ej: "driver_1" -> 1)
                            try:
                                original_driver_num = int(original_driver_id.split('_')[-1])
                            except:
                                original_driver_num = hash(original_driver_id) % 37 + 1
                        else:
                            original_driver_num = int(original_driver_id)
                        
                        # Si el conductor original está disponible, usarlo
                        if original_driver_num in available_drivers:
                            new_driver_id = original_driver_num
                        else:
                            # Buscar un reemplazo del pool de respaldo
                            # Usar round-robin entre conductores disponibles
                            if available_drivers:
                                replacement_index = (original_driver_num - 1) % len(available_drivers)
                                new_driver_id = available_drivers[replacement_index]
                            else:
                                new_driver_id = original_driver_num
                        
                        new_assignment = {
                            'date': day_date.isoformat(),
                            'week': week_number,
                            'driver_id': new_driver_id,
                            'driver_name': f'Conductor_{new_driver_id}',
                            'shift_type': pattern_assign.get('shift_type', pattern_assign.get('shift', 'T1')),
                            'start_time': pattern_assign.get('start_time', '08:00'),
                            'end_time': pattern_assign.get('end_time', '16:00'),
                            'hours': pattern_assign.get('hours', 8),
                            'service': pattern_assign.get('service', pattern_assign.get('service_name', 'Unknown')),
                            'is_replacement': new_driver_id != original_driver_num
                        }
                        
                        week_assignments.append(new_assignment)
                        annual_assignments['by_driver'][new_driver_id].append(new_assignment)
            
            annual_assignments['by_week'][week_number] = {
                'week_start': week_start.isoformat(),
                'week_end': week_end.isoformat(),
                'assignments': week_assignments,
                'drivers_active': len(available_drivers),
                'drivers_vacation': len(on_vacation),
                'total_hours': sum(a['hours'] for a in week_assignments)
            }
            
            # Avanzar a la siguiente semana
            current_date += timedelta(days=7)
            week_number += 1
        
        return annual_assignments
    
    def _calculate_annual_metrics(self, annual_assignments: Dict, 
                                 drivers: List[Dict],
                                 vacation_schedule: Dict) -> Dict:
        """Calcula métricas anuales"""
        metrics = {
            'total_weeks': len(annual_assignments['by_week']),
            'total_assignments': 0,
            'total_hours': 0,
            'hours_by_driver': {},
            'replacements_count': 0,
            'weekly_stats': []
        }
        
        # Calcular totales
        for week_num, week_data in annual_assignments['by_week'].items():
            assignments = week_data['assignments']
            metrics['total_assignments'] += len(assignments)
            metrics['total_hours'] += week_data['total_hours']
            metrics['replacements_count'] += sum(
                1 for a in assignments if a.get('is_replacement', False)
            )
            
            metrics['weekly_stats'].append({
                'week': week_num,
                'assignments': len(assignments),
                'hours': week_data['total_hours'],
                'active_drivers': week_data['drivers_active']
            })
        
        # Calcular horas por conductor
        for driver_id, assignments in annual_assignments['by_driver'].items():
            total_hours = sum(a['hours'] for a in assignments)
            metrics['hours_by_driver'][driver_id] = total_hours
        
        # Promedios
        if drivers:
            metrics['avg_hours_per_driver'] = metrics['total_hours'] / len(drivers)
            metrics['avg_weekly_hours'] = metrics['total_hours'] / 52
        
        return metrics
    
    def generate_excel_report(self, solution: Dict, client_name: str) -> str:
        """Genera reporte Excel del plan anual basado en patrón"""
        if solution['status'] != 'success':
            return None
        
        filename = f"plan_anual_patron_{client_name}_{solution['year']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Hoja 1: Resumen
            summary_data = {
                'Métrica': [
                    'Conductores Totales',
                    'Conductores Base (patrón)',
                    'Respaldo Vacaciones',
                    'Semanas Cubiertas',
                    'Total Asignaciones',
                    'Total Horas',
                    'Promedio Horas/Conductor/Año',
                    'Reemplazos por Vacaciones',
                    'Tiempo Optimización (s)'
                ],
                'Valor': [
                    solution['drivers_total'],
                    solution['drivers_base'],
                    solution['drivers_backup'],
                    solution['metrics']['total_weeks'],
                    f"{solution['metrics']['total_assignments']:,}",
                    f"{solution['metrics']['total_hours']:,.0f}",
                    f"{solution['metrics'].get('avg_hours_per_driver', 0):,.1f}",
                    solution['metrics']['replacements_count'],
                    f"{solution['optimization_time']:.2f}"
                ]
            }
            df_summary = pd.DataFrame(summary_data)
            df_summary.to_excel(writer, sheet_name='Resumen', index=False)
            
            # Hoja 2: Patrón Base (4 semanas)
            pattern_data = []
            for assignment in solution['base_pattern']['assignments'][:200]:  # Limitar para ejemplo
                pattern_data.append({
                    'Fecha': assignment['date'],
                    'Conductor ID': assignment['driver_id'],
                    'Turno': assignment['shift_type'],
                    'Inicio': assignment['start_time'],
                    'Fin': assignment['end_time'],
                    'Horas': assignment['hours'],
                    'Servicio': assignment['service']
                })
            
            if pattern_data:
                df_pattern = pd.DataFrame(pattern_data)
                df_pattern.to_excel(writer, sheet_name='Patrón 4 Semanas', index=False)
            
            # Hoja 3: Calendario de Vacaciones
            vacation_data = []
            for driver_id, vac_info in solution['vacation_schedule'].items():
                for week in vac_info['weeks']:
                    vacation_data.append({
                        'Conductor ID': driver_id,
                        'Conductor': vac_info['driver_name'],
                        'Semana': week,
                        'Período': f'Semana {week}'
                    })
            
            df_vacation = pd.DataFrame(vacation_data)
            df_vacation = df_vacation.sort_values(['Semana', 'Conductor ID'])
            df_vacation.to_excel(writer, sheet_name='Vacaciones', index=False)
            
            # Hoja 4: Resumen Semanal
            weekly_data = []
            for week_stat in solution['metrics']['weekly_stats']:
                weekly_data.append({
                    'Semana': week_stat['week'],
                    'Asignaciones': week_stat['assignments'],
                    'Horas Totales': week_stat['hours'],
                    'Conductores Activos': week_stat['active_drivers']
                })
            
            df_weekly = pd.DataFrame(weekly_data)
            df_weekly.to_excel(writer, sheet_name='Resumen Semanal', index=False)
            
            # Hoja 5: Horas por Conductor
            driver_hours_data = []
            for driver_id, hours in solution['metrics']['hours_by_driver'].items():
                driver_hours_data.append({
                    'Conductor ID': driver_id,
                    'Conductor': f'Conductor_{driver_id}',
                    'Horas Anuales': hours,
                    'Promedio Mensual': hours / 12,
                    'Promedio Semanal': hours / 52
                })
            
            df_driver_hours = pd.DataFrame(driver_hours_data)
            df_driver_hours = df_driver_hours.sort_values('Conductor ID')
            df_driver_hours.to_excel(writer, sheet_name='Horas por Conductor', index=False)
        
        # Aplicar formato
        wb = openpyxl.load_workbook(filename)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Formato de encabezados
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Ajustar anchos de columna
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(50, (max_length + 2) * 1.2)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(filename)
        return filename