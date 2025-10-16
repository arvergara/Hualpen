"""
Generador mejorado de reportes con vista detallada por conductor
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from typing import Dict, List, Any
import os


class EnhancedOutputGenerator:
    """Generador mejorado de reportes con vista detallada por conductor"""
    
    def __init__(self, solution: Dict[str, Any], client_name: str):
        self.solution = solution
        self.client_name = client_name
        self.assignments = solution.get('assignments', [])
        self.driver_stats = solution.get('driver_stats', {})
        self.metrics = solution.get('metrics', {})
        self.min_rest_between_shifts = self._resolve_min_rest_between_shifts()

    def _resolve_min_rest_between_shifts(self) -> float:
        """Derive the minimum rest requirement from the solution payload (default 10h)."""
        constraints = self.solution.get('constraints', {}) if self.solution else {}
        rest_hours = constraints.get('min_rest_between_shifts')

        if rest_hours is None:
            parameters = self.solution.get('parameters', {}) if self.solution else {}
            rest_hours = parameters.get('min_rest_hours')

        try:
            return float(rest_hours) if rest_hours is not None else 10.0
        except (TypeError, ValueError):
            return 10.0
    
    def generate_excel_with_driver_details(self) -> str:
        """Genera Excel con vista detallada por conductor"""
        wb = Workbook()
        
        # Eliminar hoja por defecto
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Crear hojas
        self._create_assignments_sheet(wb)
        self._create_driver_detail_sheet(wb)  # Nueva hoja mejorada
        self._create_summary_sheet(wb)
        
        # Guardar archivo
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"roster_{self.client_name}_{timestamp}_enhanced.xlsx"
        wb.save(filename)
        
        return filename
    
    def _create_assignments_sheet(self, wb):
        """Crea hoja de asignaciones"""
        ws = wb.create_sheet("Asignaciones")

        # Encabezados
        headers = ['Fecha', 'Día', 'Servicio', 'Turno', 'Vehículo / Tipo', 'Conductor', 'Hora Inicio', 'Hora Fin', 'Horas']
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        # Datos
        row = 2
        for assignment in sorted(self.assignments, key=lambda x: (x['date'], x['start_time'])):
            date_obj = datetime.fromisoformat(assignment['date'])
            day_names = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']
            
            ws.cell(row=row, column=1, value=assignment['date'])
            ws.cell(row=row, column=2, value=day_names[date_obj.weekday()])
            ws.cell(row=row, column=3, value=assignment.get('service_name', assignment.get('service', '')))
            ws.cell(row=row, column=4, value=f"Turno {assignment.get('shift', assignment.get('shift_number', ''))}")
            ws.cell(row=row, column=5, value=self._format_vehicle_label(assignment))
            ws.cell(row=row, column=6, value=assignment['driver_name'])
            ws.cell(row=row, column=7, value=assignment['start_time'])
            ws.cell(row=row, column=8, value=assignment['end_time'])
            
            # Calcular duration_hours si no está presente o es 0
            duration = assignment.get('duration_hours', 0)
            if duration == 0 and assignment.get('start_time') and assignment.get('end_time'):
                try:
                    start = datetime.strptime(assignment['start_time'], '%H:%M')
                    end = datetime.strptime(assignment['end_time'], '%H:%M')
                    # Si el turno cruza la medianoche
                    if end < start:
                        delta = (end.hour + 24 - start.hour) + (end.minute - start.minute) / 60
                    else:
                        delta = (end.hour - start.hour) + (end.minute - start.minute) / 60
                    duration = round(delta, 2)
                except:
                    duration = 0
            
            ws.cell(row=row, column=9, value=duration)
            
            # Resaltar domingos
            if date_obj.weekday() == 6:
                for col in range(1, 10):
                    ws.cell(row=row, column=col).fill = PatternFill(
                        start_color="FFE6E6", end_color="FFE6E6", fill_type="solid"
                    )
            
            row += 1
        
        # Ajustar anchos
        for col in range(1, 10):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _format_vehicle_label(self, assignment: Dict[str, Any]) -> str:
        """Devuelve una etiqueta amigable con el tipo de vehículo"""
        vehicle_type = assignment.get('vehicle_type') or assignment.get('vehicle_category') or ''
        vehicle_value = assignment.get('vehicle')

        type_label = str(vehicle_type).strip()
        if type_label.lower() in {'', 'unknown', 'industrial', 'urbano', 'interurbano'}:
            type_label = ''
        else:
            type_label = type_label.replace('_', ' ').title()

        idx_label = ''
        if isinstance(vehicle_value, (int, float)):
            try:
                idx_label = f"#{int(vehicle_value) + 1}"
            except (TypeError, ValueError):
                idx_label = f"{vehicle_value}"
        elif vehicle_value:
            idx_label = str(vehicle_value).strip()

        parts = [part for part in [type_label, idx_label] if part]
        if parts:
            return ' '.join(parts)
        return 'Vehículo'

    def _create_driver_detail_sheet(self, wb):
        """Crea hoja con vista detallada por conductor ordenada cronológicamente"""
        ws = wb.create_sheet("Vista por Conductor")
        
        # Título
        ws['A1'] = "DETALLE DE TURNOS POR CONDUCTOR"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:K1')
        
        # Agrupar asignaciones por conductor
        driver_assignments = {}
        for assignment in self.assignments:
            driver_name = assignment['driver_name']
            if driver_name not in driver_assignments:
                driver_assignments[driver_name] = []
            driver_assignments[driver_name].append(assignment)
        
        # Ordenar conductores
        sorted_drivers = sorted(driver_assignments.keys())
        
        row = 3
        
        for driver_name in sorted_drivers:
            # Encabezado del conductor
            ws.cell(row=row, column=1, value=f"CONDUCTOR: {driver_name}")
            ws.cell(row=row, column=1).font = Font(size=12, bold=True)
            ws.merge_cells(f'A{row}:K{row}')
            
            # Fondo para el nombre del conductor
            for col in range(1, 12):
                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"
                )
            
            row += 1
            
            # Encabezados de columnas para este conductor
            headers = ['Fecha', 'Día', 'Servicio', 'Turno', 'Inicio', 'Fin', 
                      'Horas', 'Tiempo Manejo', 'Descanso Previo', 'Jornada Total', 'Observaciones']
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
            
            row += 1
            
            # Ordenar turnos del conductor cronológicamente
            driver_shifts = sorted(driver_assignments[driver_name], 
                                 key=lambda x: (x['date'], x['start_time']))
            
            # Variables para tracking
            previous_end = None
            previous_date = None
            daily_hours = {}
            weekly_hours = 0
            sunday_count = 0
            
            for i, shift in enumerate(driver_shifts):
                date_obj = datetime.fromisoformat(shift['date'])
                day_names = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']
                day_name = day_names[date_obj.weekday()]
                
                # Calcular descanso desde turno anterior
                rest_hours = "N/A"
                if previous_end and previous_date:
                    # Crear objetos datetime para comparación
                    current_start = datetime.combine(date_obj, 
                                                    datetime.strptime(shift['start_time'], '%H:%M').time())
                    
                    # Si es el mismo día, es una pausa intra-turno
                    if previous_date == date_obj:
                        # Calcular pausa entre turnos del mismo día
                        current_time = datetime.strptime(shift['start_time'], '%H:%M')
                        pause_minutes = (current_time.hour * 60 + current_time.minute) - \
                                      (previous_end.hour * 60 + previous_end.minute)
                        pause_hours = pause_minutes / 60
                        rest_hours = f"Pausa {pause_hours:.1f}h"
                    else:
                        # Es descanso entre días diferentes
                        # Si el turno anterior terminó después de medianoche
                        if previous_end.hour < 6:  # Asumimos que turnos después de medianoche son del día anterior
                            previous_end_full = datetime.combine(previous_date + timedelta(days=1), 
                                                                previous_end.time())
                        else:
                            previous_end_full = datetime.combine(previous_date, previous_end.time())
                        
                        rest_delta = current_start - previous_end_full
                        rest_hours = round(rest_delta.total_seconds() / 3600, 1)
                
                # Calcular jornada del día (todos los turnos del mismo día)
                if shift['date'] not in daily_hours:
                    # Calcular total de horas para este día
                    day_total = sum(s.get('duration_hours', 0) 
                                  for s in driver_shifts if s['date'] == shift['date'])
                    daily_hours[shift['date']] = day_total
                
                # Tiempo de manejo (horas del turno)
                driving_hours = shift.get('duration_hours', 0)
                
                # Verificar restricciones
                observations = []
                
                # Verificar descanso mínimo (solo si es descanso entre días)
                if isinstance(rest_hours, (int, float)) and rest_hours < self.min_rest_between_shifts:
                    threshold_str = f"{self.min_rest_between_shifts:g}"
                    observations.append(f"⚠️ Descanso < {threshold_str}h ({rest_hours}h)")
                # Para pausas intra-turno, verificar proporcionalidad (2h por cada 5h de conducción)
                elif isinstance(rest_hours, str) and rest_hours.startswith("Pausa"):
                    # Extraer horas de pausa
                    pause_value = float(rest_hours.replace("Pausa ", "").replace("h", ""))
                    # El turno anterior fue de 3 horas, necesita pausa proporcional
                    required_pause = (3 / 5) * 2  # 1.2 horas
                    if pause_value < required_pause:
                        observations.append(f"⚠️ Pausa insuficiente (mínimo {required_pause:.1f}h)")
                
                # Verificar jornada diaria
                if daily_hours[shift['date']] > 12:
                    observations.append(f"⚠️ Jornada > 12h")
                
                # Contar domingos
                if date_obj.weekday() == 6:
                    sunday_count += 1
                    if sunday_count > 2:
                        observations.append("⚠️ Más de 2 domingos")
                
                # Escribir fila
                ws.cell(row=row, column=1, value=shift['date'])
                ws.cell(row=row, column=2, value=day_name)
                ws.cell(row=row, column=3, value=shift.get('service_name', shift.get('service', '')))
                ws.cell(row=row, column=4, value=f"T{shift.get('shift', shift.get('shift_number', ''))}")
                ws.cell(row=row, column=5, value=shift['start_time'])
                ws.cell(row=row, column=6, value=shift['end_time'])
                ws.cell(row=row, column=7, value=driving_hours)
                ws.cell(row=row, column=8, value=driving_hours)  # Tiempo manejando = horas del turno
                ws.cell(row=row, column=9, value=rest_hours if rest_hours != "N/A" else "")
                ws.cell(row=row, column=10, value=daily_hours[shift['date']])
                ws.cell(row=row, column=11, value="; ".join(observations) if observations else "OK")
                
                # Colorear según estado
                if observations:
                    for col in range(1, 12):
                        ws.cell(row=row, column=col).fill = PatternFill(
                            start_color="FFE6E6", end_color="FFE6E6", fill_type="solid"
                        )
                elif date_obj.weekday() == 6:  # Domingo
                    for col in range(1, 12):
                        ws.cell(row=row, column=col).fill = PatternFill(
                            start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
                        )
                
                # Actualizar para siguiente iteración
                previous_end = datetime.strptime(shift['end_time'], '%H:%M')
                previous_date = date_obj
                
                row += 1
            
            # Resumen del conductor
            total_hours = sum(shift.get('duration_hours', 0) for shift in driver_shifts)
            ws.cell(row=row, column=1, value="TOTAL")
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=7, value=total_hours)
            ws.cell(row=row, column=7).font = Font(bold=True)
            
            for col in range(1, 12):
                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"
                )
            
            row += 2  # Espacio entre conductores
        
        # Ajustar anchos de columna
        column_widths = [12, 10, 15, 8, 8, 8, 8, 12, 14, 12, 25]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
    
    def _create_summary_sheet(self, wb):
        """Crea hoja de resumen"""
        ws = wb.create_sheet("Resumen")
        
        # Título
        ws['A1'] = f"RESUMEN DE OPTIMIZACIÓN - {self.client_name}"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:D1')
        
        # Métricas principales
        row = 3
        ws[f'A{row}'] = "MÉTRICAS PRINCIPALES"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 1
        
        metrics_data = [
            ("Total de Turnos", self.metrics.get('total_shifts', 0)),
            ("Total de Horas", f"{self.metrics.get('total_hours', 0):.1f}"),
            ("Conductores Utilizados", self.metrics.get('drivers_used', 0)),
            ("Horas Promedio por Conductor", f"{self.metrics.get('avg_hours_per_driver', 0):.1f}"),
            ("Tiempo de Optimización", f"{self.metrics.get('optimization_time', 0):.2f}s"),
            ("Costo Total Estimado", f"${self.metrics.get('total_cost', 0):,.0f}")
        ]
        
        for metric_name, metric_value in metrics_data:
            ws[f'A{row}'] = metric_name
            ws[f'B{row}'] = metric_value
            row += 1
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20


def generate_enhanced_report(solution, client_name):
    """Función de conveniencia para generar reporte mejorado"""
    generator = EnhancedOutputGenerator(solution, client_name)
    return generator.generate_excel_with_driver_details()
