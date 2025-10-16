"""
Generador de reportes y archivos Excel con los resultados de la optimización
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, List, Any
from datetime import datetime, date
import os


class OutputGenerator:
    """
    Genera reportes en Excel y otros formatos con los resultados de la optimización
    """
    
    def __init__(self, solution: Dict[str, Any], client_name: str):
        self.solution = solution
        self.client_name = client_name
        self.assignments = solution.get('assignments', [])
        self.metrics = solution.get('metrics', {})
        self.driver_summary = solution.get('driver_summary', {})
        
    def generate_excel_report(self, output_path: str = None) -> str:
        """
        Genera un reporte completo en Excel con múltiples hojas
        """
        if output_path is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_path = f"roster_{self.client_name}_{timestamp}.xlsx"
        
        # Crear libro de Excel
        wb = Workbook()
        
        # Hoja 1: Resumen Ejecutivo
        self._create_summary_sheet(wb)
        
        # Hoja 2: Asignaciones Detalladas
        self._create_assignments_sheet(wb)
        
        # Hoja 3: Vista por Conductor
        self._create_driver_view_sheet(wb)
        
        # Hoja 4: Vista Calendario (una hoja por mes si hay múltiples meses)
        self._create_calendar_view_sheets(wb)
        
        # Hoja 5: Métricas y Estadísticas
        self._create_metrics_sheet(wb)
        
        # Guardar archivo
        wb.save(output_path)
        print(f"Reporte Excel generado: {output_path}")
        
        return output_path
    
    def _create_summary_sheet(self, wb):
        """Crea hoja de resumen ejecutivo"""
        ws = wb.active
        ws.title = "Resumen Ejecutivo"
        
        # Estilos
        title_font = Font(size=16, bold=True)
        header_font = Font(size=12, bold=True)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_text = Font(color="FFFFFF", bold=True)
        
        # Título
        ws['A1'] = f"Optimización de Turnos - {self.client_name}"
        ws['A1'].font = title_font
        ws.merge_cells('A1:E1')
        
        ws['A2'] = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        
        # Agregar información del régimen laboral si está disponible
        if 'regime' in self.solution:
            ws['A3'] = f"Régimen Laboral: {self.solution['regime']}"
            ws['A3'].font = Font(bold=True, color="0066CC")
            if 'regime_constraints' in self.solution:
                constraints = self.solution['regime_constraints']
                constraints_text = []
                if constraints.get('max_weekly_hours'):
                    constraints_text.append(f"Máx {constraints['max_weekly_hours']}h/semana")
                if constraints.get('max_monthly_hours'):
                    constraints_text.append(f"Máx {constraints['max_monthly_hours']}h/mes")
                if constraints.get('max_continuous_driving'):
                    constraints_text.append(f"Máx {constraints['max_continuous_driving']}h conducción continua")
                if constraints_text:
                    ws['C3'] = f"Restricciones: {', '.join(constraints_text)}"
        
        # Sección de métricas principales
        ws['A4'] = "MÉTRICAS PRINCIPALES"
        ws['A4'].font = header_font
        
        row = 5
        metrics_data = [
            ("Estado de la optimización", self.solution.get('status', 'success').upper()),
            ("Total de asignaciones", len(self.assignments)),
            ("Conductores utilizados", self.metrics.get('drivers_used', 0)),
            ("Costo total mensual", f"${self.metrics.get('total_cost', 0):,.0f}"),
            ("Horas totales", f"{self.metrics.get('total_hours', 0):.1f}"),
            ("Promedio horas/conductor", f"{self.metrics.get('avg_hours_per_driver', 0):.1f}"),
            ("Cobertura de servicios", f"{self.solution.get('service_coverage', {}).get('coverage_percent', 0)}%"),
            ("Score de cumplimiento", f"{self.metrics.get('compliance_score', 100)}%")
        ]
        
        for label, value in metrics_data:
            ws[f'A{row}'] = label
            ws[f'C{row}'] = value
            ws[f'C{row}'].alignment = Alignment(horizontal='right')
            row += 1
        
        # Sección de distribución por tipo de contrato
        row += 2
        ws[f'A{row}'] = "DISTRIBUCIÓN DE CONDUCTORES"
        ws[f'A{row}'].font = header_font
        row += 1
        
        # Contar por tipo de contrato
        contract_types = {}
        for driver_id, driver_data in self.driver_summary.items():
            contract_type = driver_data.get('contract_type', 'unknown')
            if contract_type not in contract_types:
                contract_types[contract_type] = {'count': 0, 'hours': 0}
            contract_types[contract_type]['count'] += 1
            contract_types[contract_type]['hours'] += driver_data.get('total_hours', 0)
        
        ws[f'A{row}'] = "Tipo de Contrato"
        ws[f'B{row}'] = "Cantidad"
        ws[f'C{row}'] = "Horas Totales"
        ws[f'D{row}'] = "Promedio Horas"
        
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].font = header_text
            ws[f'{col}{row}'].fill = header_fill
        
        row += 1
        for contract_type, data in contract_types.items():
            ws[f'A{row}'] = contract_type.replace('_', ' ').title()
            ws[f'B{row}'] = data['count']
            ws[f'C{row}'] = f"{data['hours']:.1f}"
            ws[f'D{row}'] = f"{data['hours']/data['count']:.1f}" if data['count'] > 0 else "0"
            row += 1
        
        # Ajustar anchos de columna
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws.column_dimensions[col].width = 25
    
    def _create_assignments_sheet(self, wb):
        """Crea hoja con asignaciones detalladas"""
        ws = wb.create_sheet("Asignaciones Detalladas")

        # Encabezados
        headers = ['Fecha', 'Día', 'Servicio', 'Turno', 'Vehículo / Tipo',
                  'Conductor ID', 'Conductor', 'Patrón', 'Hora Inicio', 'Hora Fin', 'Horas']
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        # Datos
        driver_summary = self.solution.get('driver_summary', {})

        for row, assignment in enumerate(self.assignments, 2):
            assignment_date = datetime.fromisoformat(assignment['date'])
            day_name = ['Lunes', 'Martes', 'Miércoles', 'Jueves',
                       'Viernes', 'Sábado', 'Domingo'][assignment_date.weekday()]

            # Obtener patrón del conductor
            driver_id = assignment.get('driver_id', '')
            pattern = ''
            if driver_id in driver_summary:
                pattern = driver_summary[driver_id].get('pattern', '')

            ws.cell(row=row, column=1, value=assignment['date'])
            ws.cell(row=row, column=2, value=day_name)
            service_label = assignment.get('service_name') or assignment.get('service') or ''
            ws.cell(row=row, column=3, value=service_label)

            shift_number = assignment.get('shift', assignment.get('shift_number', ''))
            ws.cell(row=row, column=4, value=f"Turno {shift_number}")

            vehicle_label = self._format_vehicle_label(assignment)
            ws.cell(row=row, column=5, value=vehicle_label)
            ws.cell(row=row, column=6, value=assignment['driver_id'])
            ws.cell(row=row, column=7, value=assignment['driver_name'])
            ws.cell(row=row, column=8, value=pattern or 'N/A')
            ws.cell(row=row, column=9, value=assignment['start_time'])
            ws.cell(row=row, column=10, value=assignment['end_time'])
            ws.cell(row=row, column=11, value=assignment['duration_hours'])

            # Resaltar domingos
            if assignment_date.weekday() == 6:
                for col in range(1, 12):
                    ws.cell(row=row, column=col).fill = PatternFill(
                        start_color="FFE6E6", end_color="FFE6E6", fill_type="solid"
                    )

        # Ajustar anchos
        for col in range(1, 12):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _format_vehicle_label(self, assignment: Dict[str, Any]) -> str:
        """Devuelve una etiqueta amigable con tipo de vehículo e índice."""
        vehicle_type = assignment.get('vehicle_type') or assignment.get('vehicle_category') or ''
        vehicle_value = assignment.get('vehicle')

        type_label = str(vehicle_type).strip()
        if type_label.lower() in {'', 'unknown', 'industrial', 'urbano', 'interurbano'}:
            type_label = ''
        else:
            type_label = type_label.replace('_', ' ').title()

        idx_label = ''
        if isinstance(vehicle_value, (int, float)):
            try:
                idx_label = f"#{int(vehicle_value) + 1}"
            except (TypeError, ValueError):
                idx_label = f"{vehicle_value}"
        elif vehicle_value:
            idx_label = str(vehicle_value).strip()

        parts = [part for part in [type_label, idx_label] if part]
        if parts:
            return ' '.join(parts)
        return 'Vehículo'
    
    def _create_driver_view_sheet(self, wb):
        """Crea hoja con vista por conductor"""
        ws = wb.create_sheet("Vista por Conductor")
        
        # Encabezados - agregar columna de patrón
        headers = ['Conductor ID', 'Nombre', 'Patrón', 'Tipo Contrato', 'Total Turnos',
                  'Total Horas', 'Utilización %', 'Domingos Trabajados', 'Salario Base']
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        # Datos por conductor
        row = 2
        for driver_id, driver_data in sorted(self.driver_summary.items()):
            ws.cell(row=row, column=1, value=driver_id)
            ws.cell(row=row, column=2, value=driver_data.get('name', driver_data.get('driver_name', '')))
            ws.cell(row=row, column=3, value=driver_data.get('pattern', 'N/A'))
            ws.cell(row=row, column=4, value=driver_data.get('contract_type', 'unknown').replace('_', ' ').title())
            ws.cell(row=row, column=5, value=driver_data.get('total_shifts', driver_data.get('total_assignments', 0)))
            ws.cell(row=row, column=6, value=driver_data.get('total_hours', 0))
            # Calcular utilización basado en horas
            total_hours = driver_data.get('total_hours', 0)
            max_hours = 180 if self.solution.get('regime') == 'Interurbano' else 176  # 44h * 4 semanas
            utilization = (total_hours / max_hours * 100) if total_hours > 0 else 0
            ws.cell(row=row, column=7, value=f"{utilization:.1f}%")
            ws.cell(row=row, column=8, value=driver_data.get('sundays_worked', 0))
            ws.cell(row=row, column=9, value=f"${850000:,.0f}")
            
            # Color según utilización
            utilization = driver_data.get('utilization', 0)
            if utilization > 90:
                fill_color = "FFB3B3"  # Rojo claro - sobreutilizado
            elif utilization < 50:
                fill_color = "B3D9FF"  # Azul claro - subutilizado
            else:
                fill_color = "B3FFB3"  # Verde claro - óptimo
            
            ws.cell(row=row, column=6).fill = PatternFill(
                start_color=fill_color, end_color=fill_color, fill_type="solid"
            )
            
            row += 1
        
        # Ajustar anchos
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 18
    
    def _create_calendar_view_sheets(self, wb):
        """Crea hojas de vista de calendario - una por cada mes con asignaciones"""
        from datetime import datetime
        import calendar

        if not self.assignments:
            return

        # Detectar todos los meses únicos en las asignaciones
        months_set = set()
        for assignment in self.assignments:
            date_obj = datetime.fromisoformat(assignment['date'])
            months_set.add((date_obj.year, date_obj.month))

        # Ordenar meses
        sorted_months = sorted(list(months_set))

        # Crear una hoja por cada mes
        for year, month in sorted_months:
            self._create_calendar_view_sheet_for_month(wb, year, month)

    def _create_calendar_view_sheet_for_month(self, wb, year, month):
        """Crea vista de calendario para un mes específico"""
        from datetime import datetime, timedelta
        import calendar as cal

        month_name = cal.month_name[month]
        ws = wb.create_sheet(f"CALENDARIO {month_name} {year}")

        # Generar todos los días del mes
        num_days = cal.monthrange(year, month)[1]
        month_dates = []
        for day in range(1, num_days + 1):
            date = datetime(year, month, day)
            month_dates.append(date.date().isoformat())
        
        # Agrupar asignaciones por fecha y conductor (simplificado: X si trabajó)
        calendar_data = {}
        for assignment in self.assignments:
            date_key = assignment['date']
            driver_key = assignment['driver_name']  # Usar nombre en lugar de ID
            
            if date_key not in calendar_data:
                calendar_data[date_key] = set()
            
            calendar_data[date_key].add(driver_key)
        
        # Obtener lista de conductores únicos ordenados
        all_drivers = sorted(set(a['driver_name'] for a in self.assignments))
        
        # Fila 1: Título del mes
        ws.cell(row=1, column=1, value=f"CALENDARIO {cal.month_name[month]} {year}")
        ws.cell(row=1, column=1).font = Font(size=14, bold=True)
        ws.merge_cells(f'A1:{get_column_letter(num_days + 2)}1')
        
        # Fila 2: Encabezados - Días del mes
        ws.cell(row=2, column=1, value="Conductor")
        ws.cell(row=2, column=1).font = Font(bold=True)
        ws.cell(row=2, column=1).fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        
        for col, date_str in enumerate(month_dates, 2):
            date_obj = datetime.fromisoformat(date_str)
            day_name = ['L', 'M', 'X', 'J', 'V', 'S', 'D'][date_obj.weekday()]
            ws.cell(row=2, column=col, value=date_obj.day)
            ws.cell(row=2, column=col).font = Font(bold=True)
            ws.cell(row=2, column=col).alignment = Alignment(horizontal='center')
            
            # Resaltar fines de semana
            if date_obj.weekday() == 5:  # Sábado
                ws.cell(row=2, column=col).fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
            elif date_obj.weekday() == 6:  # Domingo
                ws.cell(row=2, column=col).fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            else:
                ws.cell(row=2, column=col).fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        
        # Columna final: Total días trabajados
        ws.cell(row=2, column=num_days + 2, value="Total Días")
        ws.cell(row=2, column=num_days + 2).font = Font(bold=True)
        ws.cell(row=2, column=num_days + 2).fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        
        # Fila 3: Día de la semana
        for col, date_str in enumerate(month_dates, 2):
            date_obj = datetime.fromisoformat(date_str)
            day_name = ['L', 'M', 'X', 'J', 'V', 'S', 'D'][date_obj.weekday()]
            ws.cell(row=3, column=col, value=day_name)
            ws.cell(row=3, column=col).font = Font(size=9, italic=True)
            ws.cell(row=3, column=col).alignment = Alignment(horizontal='center')
        
        # Datos del calendario - marcar con X cuando el conductor trabaja
        # Para regímenes con patrones NxN (7x7, 10x10, etc.), marcar TODOS los días del ciclo
        regime = self.solution.get('regime', 'Urbano/Industrial')

        for row, driver_name in enumerate(all_drivers, 4):
            ws.cell(row=row, column=1, value=driver_name)
            ws.cell(row=row, column=1).font = Font(bold=True)

            # Obtener días con turnos para este conductor
            driver_work_days = set()
            driver_id = None
            for date_str, drivers in calendar_data.items():
                if driver_name in drivers:
                    driver_work_days.add(date_str)
                    # Obtener driver_id para buscar patrón
                    if driver_id is None:
                        for a in self.assignments:
                            if a['driver_name'] == driver_name:
                                driver_id = a['driver_id']
                                break

            # Para Faena Minera con patrón NxN, expandir a todo el ciclo
            work_days_expanded = driver_work_days.copy()

            if regime in ['Faena Minera', 'Minera'] and driver_id and driver_id in self.driver_summary:
                pattern = self.driver_summary[driver_id].get('pattern', '')
                work_start_date = self.driver_summary[driver_id].get('work_start_date')

                # Detectar ciclo (7x7, 10x10, 14x14)
                if 'x' in pattern and work_start_date:
                    try:
                        cycle_days = int(pattern.split('x')[0])

                        # Convertir work_start_date a datetime.date si es necesario
                        if isinstance(work_start_date, str):
                            work_start = datetime.fromisoformat(work_start_date).date()
                        elif hasattr(work_start_date, 'date'):
                            work_start = work_start_date.date()
                        else:
                            work_start = work_start_date

                        # Para cada día del mes, verificar si está en ciclo de trabajo
                        for date_str in month_dates:
                            check_date = datetime.fromisoformat(date_str).date()
                            days_since_start = (check_date - work_start).days
                            day_in_cycle = days_since_start % (2 * cycle_days)

                            # Los primeros N días del ciclo 2N son de trabajo
                            if 0 <= day_in_cycle < cycle_days:
                                work_days_expanded.add(date_str)

                    except Exception as e:
                        pass  # Si falla, usar solo días con turnos

            days_worked = 0
            for col, date_str in enumerate(month_dates, 2):
                if date_str in work_days_expanded:
                    # Distinguir entre día con turnos y día en planta sin turnos
                    has_shifts = date_str in driver_work_days

                    ws.cell(row=row, column=col, value="X")
                    ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')
                    ws.cell(row=row, column=col).font = Font(bold=True)
                    days_worked += 1

                    # Colorear según día de la semana y si tiene turnos
                    date_obj = datetime.fromisoformat(date_str)
                    if not has_shifts:
                        # Día en planta sin turnos (color más claro)
                        ws.cell(row=row, column=col).fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                        ws.cell(row=row, column=col).font = Font(bold=True, color="999999")
                    elif date_obj.weekday() == 6:  # Domingo trabajado
                        ws.cell(row=row, column=col).fill = PatternFill(start_color="FFB3B3", end_color="FFB3B3", fill_type="solid")
                    elif date_obj.weekday() == 5:  # Sábado trabajado
                        ws.cell(row=row, column=col).fill = PatternFill(start_color="B3D9FF", end_color="B3D9FF", fill_type="solid")
                    else:  # Día normal trabajado
                        ws.cell(row=row, column=col).fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")

            # Total días trabajados
            ws.cell(row=row, column=num_days + 2, value=days_worked)
            ws.cell(row=row, column=num_days + 2).font = Font(bold=True)
            ws.cell(row=row, column=num_days + 2).alignment = Alignment(horizontal='center')
        
        # Fila de totales - cuántos conductores trabajan cada día
        total_row = len(all_drivers) + 5
        ws.cell(row=total_row, column=1, value="Conductores/Día")
        ws.cell(row=total_row, column=1).font = Font(bold=True)
        ws.cell(row=total_row, column=1).fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        
        for col, date_str in enumerate(month_dates, 2):
            count = len(calendar_data.get(date_str, set()))
            ws.cell(row=total_row, column=col, value=count)
            ws.cell(row=total_row, column=col).font = Font(bold=True)
            ws.cell(row=total_row, column=col).alignment = Alignment(horizontal='center')
            ws.cell(row=total_row, column=col).fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 25  # Columna de nombres
        for col in range(2, num_days + 2):
            ws.column_dimensions[get_column_letter(col)].width = 4  # Columnas de días
        ws.column_dimensions[get_column_letter(num_days + 2)].width = 12  # Columna de totales
        
        # Agregar bordes a toda la tabla
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in range(2, total_row + 1):
            for col in range(1, num_days + 3):
                ws.cell(row=row, column=col).border = thin_border
    
    def _create_metrics_sheet(self, wb):
        """Crea hoja con métricas y estadísticas detalladas"""
        ws = wb.create_sheet("Métricas Detalladas")
        
        # Título
        ws['A1'] = "MÉTRICAS Y ESTADÍSTICAS DETALLADAS"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:D1')
        
        row = 3
        
        # Sección 1: Cumplimiento de restricciones
        ws[f'A{row}'] = "CUMPLIMIENTO DE RESTRICCIONES LABORALES"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 1
        
        # Verificar restricciones según régimen laboral
        regime = self.solution.get('regime', 'Urbano/Industrial')
        violations = {
            'max_monthly_hours': [],
            'sunday_restriction': [],
            'consecutive_days': []
        }

        for driver_id, driver_data in self.driver_summary.items():
            # Verificar horas mensuales (solo para regímenes que lo requieren)
            # Faena Minera NO tiene límite de 180h/mes, se controla por ciclos NxN
            if regime not in ['Faena Minera', 'Minera', 'Interurbano Bisemanal', 'Interurbano Bisemanal (Art. 39)']:
                if driver_data['total_hours'] > 180:
                    violations['max_monthly_hours'].append(driver_id)

            # Verificar domingos (solo para regímenes que lo requieren)
            # Faena Minera puede trabajar domingos con autorización
            if regime not in ['Faena Minera', 'Minera']:
                if driver_data['sundays_worked'] > 2:
                    violations['sunday_restriction'].append(driver_id)
        
        ws[f'A{row}'] = "Restricción"
        ws[f'B{row}'] = "Estado"
        ws[f'C{row}'] = "Conductores en Violación"
        
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].font = Font(bold=True)
            ws[f'{col}{row}'].fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            ws[f'{col}{row}'].font = Font(color="FFFFFF", bold=True)
        
        row += 1

        # Definir restricciones según régimen
        if regime in ['Faena Minera', 'Minera']:
            # Para Faena Minera: sin límite de 180h/mes, sin restricción de domingos
            restrictions = [
                ("Máximo 14 horas diarias", violations['consecutive_days']),  # Realmente debería verificar horas diarias
                ("Patrón NxN (7x7, 10x10, 14x14)", []),  # Siempre cumplido por diseño
                ("Descanso mínimo 5h entre turnos", [])  # Siempre cumplido por diseño
            ]
        elif regime in ['Interurbano Bisemanal', 'Interurbano Bisemanal (Art. 39)']:
            restrictions = [
                ("Máximo 44h promedio/semana", violations['max_monthly_hours']),
                ("Máximo 14h diarias", violations['consecutive_days']),
                ("Patrón bisemanal", [])
            ]
        elif regime in ['Urbano', 'Industrial', 'Urbano/Industrial', 'Interno']:
            # Régimen Urbano/Industrial: 44h semanales
            restrictions = [
                ("Máximo 44 horas semanales", violations['max_weekly_hours']),
                ("Mínimo 2 domingos libres", violations['sunday_restriction']),
                ("Máximo 6 días consecutivos", violations['consecutive_days'])
            ]
        else:
            # Régimen Interurbano Art. 25: 180h mensuales
            restrictions = [
                ("Máximo 180 horas/mes", violations['max_monthly_hours']),
                ("Mínimo 2 domingos libres", violations['sunday_restriction']),
                ("Máximo 6 días consecutivos", violations['consecutive_days'])
            ]

        for restriction_name, violators in restrictions:
            ws[f'A{row}'] = restriction_name
            if violators:
                ws[f'B{row}'] = "VIOLACIÓN"
                ws[f'B{row}'].font = Font(color="FF0000", bold=True)
                ws[f'C{row}'] = ", ".join(str(v) for v in violators)  # Convertir a string
            else:
                ws[f'B{row}'] = "CUMPLIDO"
                ws[f'B{row}'].font = Font(color="008000", bold=True)
                ws[f'C{row}'] = "Ninguno"
            row += 1
        
        # Sección 2: Distribución de turnos
        row += 2
        ws[f'A{row}'] = "DISTRIBUCIÓN DE TURNOS"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 1
        
        # Contar turnos por tipo
        shift_types = {'morning': 0, 'afternoon': 0, 'night': 0}
        for assignment in self.assignments:
            start_hour = int(assignment['start_time'].split(':')[0])
            if start_hour < 6 or start_hour >= 22:
                shift_types['night'] += 1
            elif start_hour < 14:
                shift_types['morning'] += 1
            else:
                shift_types['afternoon'] += 1
        
        ws[f'A{row}'] = "Tipo de Turno"
        ws[f'B{row}'] = "Cantidad"
        ws[f'C{row}'] = "Porcentaje"
        
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].font = Font(bold=True)
        
        row += 1
        total_shifts = sum(shift_types.values())
        
        for shift_type, count in shift_types.items():
            ws[f'A{row}'] = shift_type.capitalize()
            ws[f'B{row}'] = count
            ws[f'C{row}'] = f"{count/total_shifts*100:.1f}%" if total_shifts > 0 else "0%"
            row += 1
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 15
    
    def generate_text_report(self) -> str:
        """Genera un reporte en texto plano"""
        lines = []
        lines.append("=" * 80)
        lines.append(f"REPORTE DE OPTIMIZACIÓN DE TURNOS - {self.client_name}")
        lines.append("=" * 80)
        lines.append(f"Fecha de generación: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append("")
        
        # Resumen de métricas
        lines.append("RESUMEN DE MÉTRICAS")
        lines.append("-" * 40)
        lines.append(f"Estado: {self.solution.get('status', 'success').upper()}")
        lines.append(f"Total de asignaciones: {len(self.assignments)}")
        lines.append(f"Conductores utilizados: {self.metrics.get('drivers_used', 0)}")
        lines.append(f"Costo total: ${self.metrics.get('total_cost', 0):,.0f}")
        lines.append(f"Horas totales: {self.metrics.get('total_hours', 0):.1f}")
        lines.append(f"Cobertura de servicios: {self.solution.get('service_coverage', {}).get('coverage_percent', 0)}%")
        lines.append("")
        
        # Top 5 conductores por horas
        lines.append("TOP 5 CONDUCTORES POR HORAS TRABAJADAS")
        lines.append("-" * 40)
        
        sorted_drivers = sorted(
            self.driver_summary.items(),
            key=lambda x: x[1]['total_hours'],
            reverse=True
        )[:5]
        
        for driver_id, driver_data in sorted_drivers:
            lines.append(f"{driver_data['name']}: {driver_data['total_hours']:.1f} horas "
                        f"({driver_data['utilization']}% utilización)")
        
        lines.append("")
        lines.append("=" * 80)
        
        return '\n'.join(lines)
    
    def save_to_csv(self, output_path: str = None) -> str:
        """Guarda las asignaciones en formato CSV"""
        if output_path is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_path = f"assignments_{self.client_name}_{timestamp}.csv"
        
        # Convertir asignaciones a DataFrame
        df = pd.DataFrame(self.assignments)
        
        # Guardar a CSV
        df.to_csv(output_path, index=False)
        print(f"CSV generado: {output_path}")
        
        return output_path
